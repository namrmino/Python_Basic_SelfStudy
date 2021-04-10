import openpyxl
import os 
import re

di = r'C:\Users\happy\Documents\Study'
namesRegex = re.compile(r'\\{1}')
di = namesRegex.sub('\\\\', di)
os.chdir(di)
###

wb = openpyxl.Workbook()
wb.save('test.xlsx')

#
wb = openpyxl.Workbook()
sheet = wb.active
sheet2 = wb.create_sheet('두번째 시트')
sheet2 = wb['두번째 시트']
sheet2.title = '수집 데이터'

#
wb = openpyxl.Workbook()
sheet = wb.active

sheet['B2'] = 'b2'
sheet.cell(row=3, column=3).value = '3, 3'
sheet.append([1, 2, 3, 4, 5])

wb.save('test2.xlsx')

#
wb = openpyxl.load_workbook('test2.xlsx')
sheet1 = wb.active
sheet1.title = "이름 변경"

sheet1.append(range(10))
wb.save('test2.xlsx')